'''
Created on 15 ago. 2017

@author: Val
'''
import logging, pyexcel
from openpyxl import load_workbook
from pyGenealogy.common_profile import gen_profile
from pyGenealogy.gen_utils import is_year, naming_conventions, get_children_surname, get_name_from_fullname, get_partner_gender
from pyGenealogy.gen_utils import get_name_surname_from_complete_name, get_splitted_name_from_complete_name, get_location_standard
from datetime import datetime
from messages.pyFS_messages import NO_VALID_NAMING_CONVENTION, ENDED, NO_VALID_FILE, NOT_EXISTING_FILE, NO_GENI_EXECUTION, NO_GENI_KEY
from messages.pyFS_messages import COLUMN_NOT_FOUND
from pyGeni import profile
from pyGenealogy import NOT_KNOWN_VALUE
import os

from pyGedcom.gedcom_profile import gedcom_profile
from pyGeni.geniapi_common import geni_calls
from pyGedcom.gedcom_database import db_gedcom

DIFFERNCE_BIRTH_BAPT = 90
ignored_fields =["batch_number", "score", "role_in_record", "father_full_name", "mother_full_name", "easy_unique_id", "record_url", "subcollection_id"]
date_fields = {"birth_date" : "birth" , "burial_date" : "burial", "chr_date" : "baptism",
               "residence_date" : "residence", "death_date" : "death", "marriage_date" : "marriage"}
LOCATION_EQUIVALENCE = {"birth_place_text" : "birth", "death_place_text" : "death", "residence_place_text" : "residence",
                        "chr_place_text" : "baptism", "marriage_place_text" : "marriage"}

class getFSfamily(object):
    '''
    This class reads the  FS output excel from the website of FamilySearch
    '''
    def __init__(self, filename, naming_convention = "father_surname", language = "en"):
        '''
        This contructor reads the output file from FS in xlsx format
        ''' #TODO: include further naming conventions
        self.correct_execution = True
        self.language = language
        if os.path.exists(filename):
            if (not ".xlsx" in filename):
                #This file is not in xlsx format, we change it:
                pyexcel.save_book_as(file_name=filename, dest_file_name=filename + "x")
                filename = filename + "x"
        else:
            #The file does not exists, we create an error!
            logging.error(NOT_EXISTING_FILE + filename)
            self.correct_execution = False
        if (not naming_convention in naming_conventions):
            logging.error(NO_VALID_NAMING_CONVENTION)
            self.correct_execution = False
        self.naming_convention = naming_convention
        if self.correct_execution:
            self.loaded_data = load_workbook(filename, data_only=True)
            if (self.__locate_key_fields__()):
                self.profiles = []
                self.geni_profiles = []
                self.related_profiles = {}
                self.related_geni_profiles = []
                self.parents_profiles = {}
                self.parents_geni_profiles = []
                self.__get_profiles__()
            else:
                #If no we do not locate the key files we continue
                self.correct_execution = False
                logging.error(NO_VALID_FILE)
    def __locate_key_fields__(self):
        '''
        This function is used to locate the key areas within the file in order to later
        on filter
        '''
        for sheet in self.loaded_data:
            for row in sheet.iter_rows():
                for cell in row:
                    #Ok, we are looking for the right location and usually the first line contains score
                    if(cell.internal_value == "score"):
                        self.sheet_title = sheet.title
                        self.initial_row = cell.row
                        self.initial_column = cell.column
                        return True
        return False
    #TODO: simplify including intermediate functions
    def __get_profiles__(self):
        '''
        This function will take all different profiles included inside the excel file
        '''
        current_sheet = self.loaded_data[self.sheet_title]
        #Iterator of missing inptus
        number_missing = 0
        #The id number to be used
        id_profiles = 0
        #Temporal variable checking the correct reading
        correct_introduction = True
        #Intermediate variables for potential parent surnames in the input file
        potential_father_surname = []
        potential_father_surname_repetitions = []
        potential_mother_surname = []
        potential_mother_surname_repetitions = []
        #Intermediate variables for potential parent names in the input file
        potential_father_name = []
        potential_father_name_repetitions = []
        potential_mother_name = []
        potential_mother_name_repetitions = []
        #We firstly detect the surnames of the parents of the profile,we cannot avoid the double
        #iteration
        for row in range(self.initial_row+1, self.loaded_data[self.sheet_title].max_row+1):
            for column_index in range(self.initial_column,self.loaded_data[self.sheet_title].max_column):
                column_criteria = current_sheet.cell(row=self.initial_row, column=column_index).value
                cell_value = current_sheet.cell(row=row, column=column_index).value
                if (column_criteria in ["father_full_name", "mother_full_name"]  ):
                    #If the cell_value is null we shall avoid continuing
                    if (cell_value != None):
                        name_data = get_name_surname_from_complete_name(cell_value, convention=self.naming_convention, language=self.language)
                        #We have two surnames or one?
                        surname_cand = name_data[1]
                        if (name_data[2] == 2):
                            surname_cand = name_data[1].split()[0]
                        if(column_criteria == "father_full_name"):
                            if (not surname_cand in potential_father_surname):
                                potential_father_surname.append(surname_cand)
                                potential_father_surname_repetitions.append(1)
                            else:
                                index = potential_father_surname.index(surname_cand)
                                potential_father_surname_repetitions[index] = potential_father_surname_repetitions[index] + 1
                            if (not name_data[0] in potential_father_name):
                                potential_father_name.append(name_data[0])
                                potential_father_name_repetitions.append(1)
                            else:
                                index = potential_father_name.index(name_data[0])
                                potential_father_name_repetitions[index] = potential_father_name_repetitions[index] + 1
                        elif(column_criteria == "mother_full_name"):
                            if (not surname_cand in potential_mother_surname):
                                potential_mother_surname.append(surname_cand)
                                potential_mother_surname_repetitions.append(1)
                            else:
                                index = potential_mother_surname.index(surname_cand)
                                potential_mother_surname_repetitions[index] = potential_mother_surname_repetitions[index] + 1
                            if (not name_data[0] in potential_mother_name):
                                potential_mother_name.append(name_data[0])
                                potential_mother_name_repetitions.append(1)
                            else:
                                index = potential_mother_name.index(name_data[0])
                                potential_mother_name_repetitions[index] = potential_mother_name_repetitions[index] + 1
        index_father_surname = potential_father_surname_repetitions.index(max(potential_father_surname_repetitions))
        index_mother_surname = potential_mother_surname_repetitions.index(max(potential_mother_surname_repetitions))
        father_surname = potential_father_surname[index_father_surname]
        mother_surname = potential_mother_surname[index_mother_surname]
        index_father_name = potential_father_name_repetitions.index(max(potential_father_name_repetitions))
        index_mother_name = potential_mother_name_repetitions.index(max(potential_mother_name_repetitions))
        father_name = potential_father_name[index_father_name]
        mother_name = potential_mother_name[index_mother_name]
        self.father_profile = gen_profile(father_name, father_surname)
        self.mother_profile = gen_profile(mother_name, mother_surname)
        children_surname = get_children_surname(father_surname, mother_surname, self.naming_convention)
        #Now we read the complete file
        for row in range(self.initial_row+1, self.loaded_data[self.sheet_title].max_row+1):
            included_profile = gen_profile("TBD", children_surname)
            included_right = True
            for column_index in range(self.initial_column,self.loaded_data[self.sheet_title].max_column):
                column_criteria = current_sheet.cell(row=self.initial_row, column=column_index).value
                cell_value = current_sheet.cell(row=row, column=column_index).value
                #We are ignoring all those cells that are empty.
                if ( cell_value ):
                    this_introduction = True
                    #Ok, now we go one by one each of the different values
                    if(column_criteria == "gender"):
                        this_introduction = included_profile.setCheckedGender(cell_value)
                    elif (column_criteria in LOCATION_EQUIVALENCE.keys()):
                        included_profile.setPlaces(LOCATION_EQUIVALENCE[column_criteria], cell_value, self.language)
                    elif (column_criteria == "person_url"):
                        included_profile.setWebReference("https://familysearch.org/" +cell_value)
                    elif (column_criteria in date_fields.keys()):
                        #Notice that we shall detect if the given date is a year or a specific date
                        #we will make the different using "about" and using datetime in the background
                        if(is_year(cell_value)):
                            this_introduction = self.__include_a_date__(column_criteria, included_profile, datetime.strptime(str(cell_value.replace(" ", "")), "%Y").date(), "ABOUT")
                        else:
                            this_introduction = self.__include_a_date__(column_criteria, included_profile, datetime.strptime(cell_value, "%d %b %Y").date(), "EXACT")
                    elif(column_criteria == "full_name"):
                        included_profile.set_name(get_name_from_fullname(cell_value,potential_father_surname, potential_mother_surname, language=self.language))
                        #In the case the name if not the same, we create it as nickname
                        if (cell_value != included_profile.returnFullName()): included_profile.add_nickname(cell_value)
                    elif (column_criteria == "spouse_full_name"):
                        #Here we create a new profile using the surname of the guy
                        names = get_name_surname_from_complete_name(cell_value, convention=self.naming_convention, language=self.language)
                        partner = gen_profile(names[0], names[1])
                        partner.set_id(id_profiles)
                        #If the surname is changed we shall include the previous surname in the nicknames
                        if (cell_value != partner.returnFullName()): partner.add_nickname(cell_value)
                        #Now we link the profiles
                        included_profile.set_marriage_id_link(id_profiles)
                        self.related_profiles[id_profiles] = partner
                    elif (column_criteria == "other_full_names"):
                        #The separator provided by family search is semicolumn
                        parents = cell_value.split(";")
                        #We obtain firstly the different names
                        father_name, father_surname, _ = get_name_surname_from_complete_name(parents[0], convention=self.naming_convention, language=self.language)
                        if (len(parents) == 2):
                            mother_name, mother_surname, _ = get_name_surname_from_complete_name(parents[1], convention=self.naming_convention, language=self.language)
                        #The algorithm provides an empty surname, we fill it with not known
                        if (father_surname == ""): father_surname = NOT_KNOWN_VALUE
                        if (mother_surname == ""): mother_surname = NOT_KNOWN_VALUE
                        #Create the standard profiles
                        father = gen_profile(father_name, father_surname)
                        mother = gen_profile(mother_name, mother_surname)
                        #If the surname is changed we shall include the previous surname in the nicknames
                        if (parents[0] != father.returnFullName()): father.add_nickname(parents[0])
                        if (len(parents) == 2) and (parents[1] != mother.returnFullName()): mother.add_nickname(parents[1])
                        #add gender
                        father.setCheckedGender("M")
                        mother.setCheckedGender("F")
                        self.parents_profiles[id_profiles] = [father, mother]
                    elif (column_criteria in ignored_fields):
                        pass
                    else:
                        number_missing = number_missing + 1
                        logging.warning(COLUMN_NOT_FOUND + column_criteria)
                    if (not this_introduction): included_right = False
                #This is a way to later on identify the link between the profiles
            id_profiles += 1
            if(not included_right) : correct_introduction = False
            self.profiles.append(included_profile)
        #Now we know the data we fix some with the proper logic
        for profile_obtained in self.profiles:
            #If the baptism and birth are close enough we assign the birth place to the baptism place
            birth_event = profile_obtained.gen_data.get("birth", None)
            bapt_event = profile_obtained.gen_data.get("baptism", None)
            if birth_event and bapt_event:
                difference = bapt_event.get_date() - birth_event.get_date()
                if abs(difference.days) < DIFFERNCE_BIRTH_BAPT:
                    place_birth = None
                    place_baptism = None
                    if profile_obtained.gen_data.get("birth", None) and profile_obtained.gen_data.get("birth", None).get_location():
                        place_birth = profile_obtained.gen_data.get("birth", None).get_location().get("raw", None)
                    if profile_obtained.gen_data.get("baptism", None) and profile_obtained.gen_data.get("baptism", None).get_location():
                        place_birth = profile_obtained.gen_data.get("baptism", None).get_location().get("raw", None)
                    if place_baptism and not place_birth:
                        profile_obtained.setPlaces("birth",get_location_standard(profile_obtained.gen_data["baptism"].get_location()), self.language)
            if profile_obtained.gen_data.get("marriage_link", None) in self.related_profiles.keys():
                id_of_marriage = profile_obtained.gen_data["marriage_link"]
                partner = self.related_profiles[id_of_marriage]
                partner.setWebReference(profile_obtained.gen_data["web_ref"])
                #It is a partner so we add as opposite sex!
                partner.setCheckedGender(get_partner_gender(profile_obtained.gen_data["gender"]))
                partner.setNewEvent(profile_obtained.gen_data["marriage"])
                partner.setPlaces("marriage", profile_obtained.gen_data["marriage"].get_location()["raw"], language=self.language )
                if id_of_marriage in self.parents_profiles.keys():
                    father = self.parents_profiles[id_of_marriage][0]
                    mother = self.parents_profiles[id_of_marriage][1]
                    father.setWebReference(profile_obtained.gen_data["web_ref"])
                    mother.setWebReference(profile_obtained.gen_data["web_ref"])
                    surnames = get_splitted_name_from_complete_name(partner.gen_data["surname"], language=self.language)[0]
                    if (father.gen_data["surname"] == NOT_KNOWN_VALUE):
                        #It might be the case that the surname is empty
                        #Ok the data was not including the right data, but we know the surname
                        if (self.naming_convention == "spanish_surname" and len(surnames) != 0):
                            father.gen_data["surname"] = surnames[0]
                        else:
                            father.gen_data["surname"] = partner.gen_data["surname"]
                    if (mother.gen_data["surname"] == NOT_KNOWN_VALUE) and (self.naming_convention == "spanish_surname") and (len(surnames) == 2):
                        mother.gen_data["surname"] = surnames[1]
                    if (self.naming_convention == "spanish_surname"):
                        #We need to ensure 2 surnames in spanish naming conventions
                        if not (mother.gen_data["surname"] in partner.gen_data["surname"]) or (len(partner.gen_data["surname"].split()) == 1):
                            #In the case we have 2 surnames, we try to eliminate the second surnames.
                            partner_surname_data = get_splitted_name_from_complete_name(partner.gen_data["surname"], language=self.language)
                            mother_surname_data = get_splitted_name_from_complete_name(mother.gen_data["surname"], language=self.language)
                            if len(partner.gen_data["nicknames"]) == 0: partner.add_nickname(partner.returnFullName())
                            partner.gen_data["surname"] = " ".join([partner_surname_data[0][0], mother_surname_data[0][0]])
        #Finally, let's merge those profiles that are the same!
        indexes_to_remove = []
        iterating_list = list(self.profiles)
        for i in range(len(iterating_list)):
            #We are going one by one all the different profiles
            if not i in indexes_to_remove:
                for j, other_prof in enumerate(iterating_list[i+1:]):
                    merged = self.profiles[i].merge_profile(other_prof, language=self.language, convention=self.naming_convention)
                    if merged:
                        indexes_to_remove.append(i+j+1)
        new_values = list(set(indexes_to_remove))
        new_values.sort()
        for deletion in reversed(new_values):
            del self.profiles[deletion]
        return correct_introduction
    @classmethod
    def __include_a_date__(cls, column_criteria, profile, date_object, accuracy ):
        '''
        Function to avoid repetition
        '''
        #No check if included in data_fileds as already included in the check!
        if accuracy == "ABOUT":
            return profile.setCheckedDate(date_fields[column_criteria], date_object.year, accuracy=accuracy)
        else:
            return profile.setCheckedDate(date_fields[column_criteria], date_object.year,date_object.month,date_object.day, accuracy=accuracy)
    def create_profiles_in_Geni(self, geni_data):
        '''
        This method will create the needed profiles directly in Geni
        '''
        if not self.correct_execution:
            logging.error(NO_GENI_EXECUTION)
            return False
        else:
            connector = geni_calls()
            valid = connector.check_valid_genikey()
            if not valid:
                #Ok, it appears the call is not correct and we are getting an error message
                logging.error(NO_GENI_KEY)
                return False
            else:
                parent_data = profile.profile(geni_input=geni_data)
                other_parent = None
                #We only make it for a single partner
                if len(parent_data.partner) == 1:
                    other_parent = profile.profile(geni_input=parent_data.partner[0])
                #We need to understand if it is mother of father, and then the logic diverts
                if (parent_data.gen_data["gender"] == "M"):
                    self.father_profile.merge_profile(parent_data, language = self.language, convention = self.naming_convention)
                    if other_parent: self.mother_profile.merge_profile(other_parent, language = self.language, convention = self.naming_convention)
                else:
                    self.mother_profile.merge_profile(parent_data, language = self.language, convention = self.naming_convention)
                    if other_parent: self.father_profile.merge_profile(other_parent, language = self.language, convention = self.naming_convention)
                #Once parents have been read, we proceed to create the profiles.
                for profile_obtained in self.profiles:
                    logging.info(profile_obtained.returnFullName())
                    profile.profile.create_as_a_child(profile_obtained, geni_input=geni_data )
                    self.geni_profiles.append(profile_obtained)
                    logging.info(profile_obtained.geni_specific_data["url"])
                    if profile_obtained.gen_data.get("marriage_link", None) in self.related_profiles.keys():
                        id_of_marriage = profile_obtained.gen_data["marriage_link"]
                        partner = self.related_profiles[id_of_marriage]
                        profile.profile.create_as_a_partner(partner, geni_input=profile_obtained.geni_specific_data["id"],
                                                    type_geni="" )
                        self.related_geni_profiles.append(partner)
                        logging.info(partner.geni_specific_data["url"])
                        if id_of_marriage in self.parents_profiles.keys():
                            father = self.parents_profiles[id_of_marriage][0]
                            mother = self.parents_profiles[id_of_marriage][1]
                            profile.profile.create_as_a_parent(father, geni_input=partner.geni_specific_data["id"], type_geni="" )
                            profile.profile.create_as_a_parent(mother, geni_input=partner.geni_specific_data["id"], type_geni="" )
                            self.parents_geni_profiles.append(father)
                            self.parents_geni_profiles.append(mother)
        logging.info(ENDED)
        return True
    def create_gedcom_file(self, output):
        '''
        This function will create the gedcom file that can be read by other genealogical tools
        '''
        if not self.correct_execution:
            logging.error(NO_GENI_EXECUTION)
            return False
        else:
            #We create first the database, a Gedcom one
            dbged = db_gedcom()
            #We make father and mother a gedcom_profile
            gedcom_profile.convert_gedcom(self.father_profile)
            gedcom_profile.convert_gedcom(self.mother_profile)
            #We add now the father and mother to the database
            id_father = dbged.add_profile(self.father_profile)
            id_mother = dbged.add_profile(self.mother_profile)
            #We prepare the children adding
            children_ged = []
            for profile_obtained in self.profiles:
                #We convert each profile into the gedcom profile
                profile_temp = profile_obtained
                gedcom_profile.convert_gedcom(profile_temp)
                id_child = dbged.add_profile(profile_temp)
                children_ged.append(id_child)
                if profile_obtained.gen_data.get("marriage_link", None) in self.related_profiles.keys():
                    id_of_marriage = profile_obtained.gen_data["marriage_link"]
                    #We capture the partner
                    partner = self.related_profiles[id_of_marriage]
                    gedcom_profile.convert_gedcom(partner)
                    #We add the partner
                    id_partner = dbged.add_profile(partner)
                    #We need to create a family here of both partners
                    if profile_obtained.gen_data["gender"] == "M":
                        dbged.add_family(father = id_child, mother = id_partner)
                    else:
                        dbged.add_family(father = id_partner, mother = id_child)
                    if id_of_marriage in self.parents_profiles.keys():
                        father = self.parents_profiles[id_of_marriage][0]
                        mother = self.parents_profiles[id_of_marriage][1]
                        #We now convert the profiles into the profile
                        gedcom_profile.convert_gedcom(father)
                        gedcom_profile.convert_gedcom(mother)
                        #We add them to the file
                        id_partner_father = dbged.add_profile(father)
                        id_partner_mother = dbged.add_profile(mother)
                        #And the family!!!
                        dbged.add_family(father = id_partner_father, mother = id_partner_mother, children = [id_partner])
                        
            #We create here the family
            dbged.add_family(father = id_father, mother = id_mother, children = children_ged)
            dbged.save_gedcom_file(output)